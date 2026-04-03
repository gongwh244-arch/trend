"""指数趋势强度统计工具 - 主脚本"""

import os
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from typing import Optional

import akshare as ak
import numpy as np
import pandas as pd
from tabulate import tabulate

from config import INDEX_CONFIG, MA_PERIOD, SECTOR_CONFIG, START_DATE

# 最大重试次数
MAX_RETRIES = 3
RETRY_DELAY = 2  # 秒

# 缓存目录
CACHE_DIR = os.path.join(os.path.dirname(__file__), "cache")


def fetch_with_retry(fetch_fn, retries=MAX_RETRIES):
    """带重试的数据获取"""
    for attempt in range(retries):
        try:
            return fetch_fn()
        except Exception as e:
            if attempt < retries - 1:
                time.sleep(RETRY_DELAY)
            else:
                raise e


def _cache_path(symbol: str) -> str:
    """返回当日缓存文件路径"""
    today = datetime.now().strftime("%Y%m%d")
    return os.path.join(CACHE_DIR, f"{symbol}_{today}.csv")


def fetch_index_data(cfg: dict) -> pd.DataFrame:
    """根据配置获取指数日线数据，返回统一格式 DataFrame[date, close]"""
    source = cfg["source"]
    symbol = cfg["symbol"]
    today = datetime.now().strftime("%Y%m%d")

    # 读取缓存
    cache_file = _cache_path(symbol)
    if os.path.exists(cache_file):
        df = pd.read_csv(cache_file, parse_dates=["date"])
        df["close"] = df["close"].astype(float)
        return df

    if source == "a_share_sina":
        # 新浪A股指数接口，symbol 需要 sh/sz 前缀
        df = fetch_with_retry(lambda: ak.stock_zh_index_daily(symbol=symbol))

    elif source == "a_share_em":
        # 东财A股指数接口，用于新浪不支持的指数
        df = fetch_with_retry(lambda: ak.index_zh_a_hist(
            symbol=symbol, period="daily",
            start_date=START_DATE, end_date=today
        ))
        df = df.rename(columns={"日期": "date", "收盘": "close"})

    elif source == "hk_sina":
        # 新浪港股指数接口
        df = fetch_with_retry(lambda: ak.stock_hk_index_daily_sina(symbol=symbol))

    elif source == "us_etf":
        # 美股ETF接口
        df = fetch_with_retry(lambda: ak.stock_us_daily(symbol=symbol, adjust="qfq"))

    elif source == "global_sina":
        # 新浪全球指数接口
        df = fetch_with_retry(lambda: ak.index_global_hist_sina(symbol=symbol))

    else:
        raise ValueError(f"未知的数据源类型: {source}")

    # 统一列名
    if "date" not in df.columns:
        # 可能列名是中文
        col_map = {}
        for col in df.columns:
            if col in ("日期",):
                col_map[col] = "date"
            elif col in ("收盘",):
                col_map[col] = "close"
        if col_map:
            df = df.rename(columns=col_map)

    df = df[["date", "close"]].copy()
    df["date"] = pd.to_datetime(df["date"])
    df["close"] = df["close"].astype(float)
    df = df.sort_values("date").reset_index(drop=True)

    # 只保留 START_DATE 之后的数据
    start = pd.to_datetime(START_DATE)
    df = df[df["date"] >= start].reset_index(drop=True)

    # 写入缓存
    os.makedirs(CACHE_DIR, exist_ok=True)
    df.to_csv(cache_file, index=False)

    # 网络请求后短暂等待，防止API限流
    time.sleep(0.3)

    return df


def detect_cross(df: pd.DataFrame) -> Optional[dict]:
    """
    检测最近一次收盘价穿越MA20的事件。
    返回: {cross_date, critical_value, cross_close, direction}
    """
    valid = df.dropna(subset=["ma20"]).reset_index(drop=True)
    if len(valid) < 2:
        return None

    # 判断每日 close 相对 MA20 的位置: 1=上方, -1=下方
    valid["position"] = np.where(valid["close"] > valid["ma20"], 1, -1)

    # 从最新往前找位置变化点
    for i in range(len(valid) - 1, 0, -1):
        if valid["position"].iloc[i] != valid["position"].iloc[i - 1]:
            return {
                "cross_date": valid["date"].iloc[i],
                "critical_value": round(valid["ma20"].iloc[i], 2),
                "cross_close": round(valid["close"].iloc[i], 2),
                "direction": "上穿" if valid["position"].iloc[i] == 1 else "下穿",
            }

    # 没有穿越：取第一个有效数据
    return {
        "cross_date": valid["date"].iloc[0],
        "critical_value": round(valid["ma20"].iloc[0], 2),
        "cross_close": round(valid["close"].iloc[0], 2),
        "direction": "上穿" if valid["position"].iloc[-1] == 1 else "下穿",
    }


def calc_daily_change(df: pd.DataFrame) -> float:
    """计算当日涨幅%"""
    if len(df) < 2:
        return 0.0
    prev_close = df["close"].iloc[-2]
    curr_close = df["close"].iloc[-1]
    return round((curr_close - prev_close) / prev_close * 100, 2)


def process_index(cfg: dict) -> Optional[dict]:
    """处理单个指数，返回结果字典"""
    try:
        df = fetch_index_data(cfg)
        if len(df) < MA_PERIOD + 1:
            print(f"数据不足({len(df)}条)，跳过")
            return None

        # 计算 MA20
        df["ma20"] = df["close"].rolling(MA_PERIOD).mean()

        # 检测穿越
        cross_info = detect_cross(df)
        if cross_info is None:
            print("无法检测穿越，跳过")
            return None

        # 当日数据
        latest_close = df["close"].iloc[-1]
        latest_date = df["date"].iloc[-1]
        daily_change = calc_daily_change(df)

        # 偏离率 = (收盘价 - 临界值点) / 临界值点
        critical_value = cross_info["critical_value"]
        deviation = round((latest_close - critical_value) / critical_value * 100, 2)

        # 持续天数 = 穿越日期到最新日期的交易日数
        cross_date = cross_info["cross_date"]
        duration_days = len(df[(df["date"] >= cross_date) & (df["date"] <= latest_date)]) - 1

        # MA20斜率% = (今日MA20 - 5日前MA20) / 5日前MA20 × 100
        valid_ma = df.dropna(subset=["ma20"])
        if len(valid_ma) >= 6:
            ma20_today = valid_ma["ma20"].iloc[-1]
            ma20_5ago = valid_ma["ma20"].iloc[-6]
            ma20_slope = round((ma20_today - ma20_5ago) / ma20_5ago * 100, 2)
        else:
            ma20_slope = 0.0

        # 状态
        status = "YES" if deviation > 0 else "NO"

        return {
            "指数名称": cfg["name"],
            "状态": status,
            "当日涨幅%": daily_change,
            "收盘点位": round(latest_close, 2),
            "临界值点": critical_value,
            "偏离率%": deviation,
            "穿越日期": cross_info["cross_date"].strftime("%Y-%m-%d"),
            "持续天数": duration_days,
            "MA20斜率%": ma20_slope,
            "趋势方向": cross_info["direction"],
            "数据日期": latest_date.strftime("%Y-%m-%d"),
        }
    except Exception as e:
        print(f"获取失败: {e}")
        return None


def _write_block(ws, start_row: int, display_df: pd.DataFrame,
                 export_cols: list, title: str):
    """在 Sheet 中从 start_row 开始写入一个独立表格块（标题+表头+数据），返回下一个可用行号"""
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    last_col_letter = get_column_letter(len(export_cols))

    # 标题行（合并单元格）
    ws.merge_cells(f"A{start_row}:{last_col_letter}{start_row}")
    title_cell = ws.cell(row=start_row, column=1)
    title_cell.value = title
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[start_row].height = 36

    # 特殊着色列索引
    pct_col_names = {"当日涨幅%", "偏离率%", "MA20斜率%"}
    pct_col_indices = [i + 1 for i, col in enumerate(export_cols) if col in pct_col_names]
    status_col_idx = next((i + 1 for i, col in enumerate(export_cols) if col == "状态"), None)

    base_size = 13
    red_font = Font(color="FF0000", size=base_size)
    green_font = Font(color="008000", size=base_size)
    normal_font = Font(size=base_size)

    # 表头行
    header_row = start_row + 1
    header_font = Font(bold=True, size=base_size)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    for col_idx, col_name in enumerate(export_cols, 1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.value = col_name
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[header_row].height = 28

    center_align = Alignment(horizontal="center", vertical="center")

    # 数据行
    data_start = start_row + 2
    for df_row_idx, (_, row_data) in enumerate(display_df.iterrows()):
        row_num = data_start + df_row_idx
        for col_idx, col_name in enumerate(export_cols, 1):
            cell = ws.cell(row=row_num, column=col_idx)
            cell.value = row_data[col_name]
            cell.alignment = center_align
            cell.font = normal_font
            if col_idx == status_col_idx:
                cell.font = red_font if cell.value == "YES" else green_font
            elif col_idx in pct_col_indices:
                val = cell.value
                if val is not None and val != 0:
                    cell.font = red_font if val > 0 else green_font
        ws.row_dimensions[row_num].height = 24

    # 返回下一个可用行（数据末尾 + 1行空行间隔）
    return data_start + len(display_df) + 1


def _adjust_column_widths(ws, export_cols: list, max_row: int):
    """根据所有内容自适应列宽（跳过合并单元格的标题行）"""
    from openpyxl.utils import get_column_letter

    # 收集所有合并单元格所在的行号，这些行不参与列宽计算
    merged_rows = set()
    for merged_range in ws.merged_cells.ranges:
        for r in range(merged_range.min_row, merged_range.max_row + 1):
            merged_rows.add(r)

    for col_idx in range(1, len(export_cols) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for r in range(1, max_row + 1):
            if r in merged_rows:
                continue
            val = str(ws.cell(row=r, column=col_idx).value or "")
            char_len = sum(2 if ord(c) > 127 else 1 for c in val)
            max_len = max(max_len, char_len)
        ws.column_dimensions[col_letter].width = max_len + 4


def _print_table(display_df: pd.DataFrame, label: str, data_date: str):
    """终端打印单张表格"""
    yes_count = (display_df["状态"] == "YES").sum()
    total = len(display_df)
    table = tabulate(display_df, headers="keys", tablefmt="simple",
                     showindex=False, numalign="right", stralign="center")
    print(f"\n{'=' * 100}")
    print(f"  {label}  |  数据日期: {data_date}  |  MA周期: {MA_PERIOD}")
    print(f"{'=' * 100}")
    print(table)
    print(f"{'=' * 100}")
    print(f"  多头(YES): {yes_count}/{total}  |  空头(NO): {total - yes_count}/{total}\n")


def export_excel(index_df: pd.DataFrame, sector_df: pd.DataFrame,
                 output_dir: str = "output"):
    """导出指数和板块到同一个 Sheet，上下独立排列"""
    os.makedirs(output_dir, exist_ok=True)
    today_str = datetime.now().strftime("%Y%m%d")
    filepath = os.path.join(output_dir, f"trend_{today_str}.xlsx")

    export_cols = ["排名", "指数名称", "状态", "当日涨幅%", "收盘点位",
                   "临界值点", "偏离率%", "穿越日期", "持续天数", "MA20斜率%"]

    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "趋势总览"

    next_row = 1

    # 指数表格块
    if not index_df.empty:
        idx_display = index_df[export_cols]
        data_date = index_df["数据日期"].iloc[0]
        yes_count = (index_df["状态"] == "YES").sum()
        total = len(index_df)
        title = f"指数趋势强度统计  |  数据日期: {data_date}  |  MA周期: {MA_PERIOD}  |  多头: {yes_count}/{total}  空头: {total - yes_count}/{total}"
        next_row = _write_block(ws, next_row, idx_display, export_cols, title)
        _print_table(idx_display, "指数趋势强度统计", data_date)

    # 板块表格块
    if not sector_df.empty:
        sec_display = sector_df[export_cols]
        data_date = sector_df["数据日期"].iloc[0]
        yes_count = (sector_df["状态"] == "YES").sum()
        total = len(sector_df)
        title = f"板块趋势强度统计  |  数据日期: {data_date}  |  MA周期: {MA_PERIOD}  |  多头: {yes_count}/{total}  空头: {total - yes_count}/{total}"
        next_row = _write_block(ws, next_row, sec_display, export_cols, title)
        _print_table(sec_display, "板块趋势强度统计", data_date)

    # 统一调整列宽
    _adjust_column_widths(ws, export_cols, next_row - 1)

    wb.save(filepath)
    print(f"  已导出: {filepath}")


def _fetch_group(config_list: list, label: str) -> pd.DataFrame:
    """获取一组指数/板块数据，多线程并发处理，返回排序后的 DataFrame"""
    print(f"正在获取{label}数据...\n")
    results = []
    total = len(config_list)

    with ThreadPoolExecutor(max_workers=4) as executor:
        future_to_cfg = {executor.submit(process_index, cfg): cfg for cfg in config_list}
        for i, future in enumerate(as_completed(future_to_cfg), 1):
            cfg = future_to_cfg[future]
            try:
                result = future.result()
                if result:
                    results.append(result)
                    print(f"  [{i}/{total}] {cfg['name']}... OK  收盘:{result['收盘点位']}  偏离率:{result['偏离率%']}%")
                else:
                    print(f"  [{i}/{total}] {cfg['name']}... 跳过")
            except Exception as e:
                print(f"  [{i}/{total}] {cfg['name']}... 失败: {e}")

    if not results:
        print(f"\n所有{label}获取失败。")
        return pd.DataFrame()

    df = pd.DataFrame(results)
    df = df.sort_values("当日涨幅%", ascending=False).reset_index(drop=True)
    df.insert(0, "排名", range(1, len(df) + 1))
    return df


def main():
    index_df = _fetch_group(INDEX_CONFIG, "指数")
    sector_df = _fetch_group(SECTOR_CONFIG, "板块")

    if index_df.empty and sector_df.empty:
        print("\n所有数据获取失败，请检查网络连接。")
        return

    export_excel(index_df, sector_df)


if __name__ == "__main__":
    main()
